from openpyxl import Workbook, load_workbook
import decimal
from datetime import datetime

# DBからデータを読み取ってExcelファイルに書き出す場合、
# セルの書式など気にする必要が無ければ、pandasの方が便利、より少ない行で実装できる

wb = Workbook()
ws = wb.active
ws2 = wb.create_sheet("Mysheet", 0) # insert at first position

# 日本語特に問題なし
ws2.cell(row=1, column=1, value="あいうえお")
ws2.cell(row=1, column=2, value=100)
ws2.cell(row=1, column=3, value=1.9999999999)
ws2.cell(row=1, column=4, value=0.25)
ws2.cell(row=1, column=5, value=decimal.Decimal("2.9999999999"))
c = ws2.cell(row=1, column=6, value=datetime.now())
print(c.number_format)
# 日付の書式を設定できる（月がゼロ埋めされないのはExcelの仕様か、2021/3/13のとき）
c.number_format = 'YYYY/MM/DD H:MM:SS'
# これは文字列になる
ws2.cell(row=1, column=7, value="2021/03/13 13:11:59")

# このExcelを既に開いているとエラーになるので注意
wb.save(filename = "write_demo.xlsx")

# テンプレートとなるExcelを読み込んでそれを編集し、別ファイルとして保存もできる
wb2 = load_workbook(filename = 'template.xlsx')
ws3 = wb2['Mysheet']
ws3.cell(row=7, column=1, value="あいうえお")
ws3.cell(row=7, column=2, value=100)
ws3.cell(row=7, column=3, value=1.9999999999)
ws3.cell(row=7, column=4, value=0.25)
ws3.cell(row=7, column=5, value=decimal.Decimal("2.9999999999"))
c = ws3.cell(row=7, column=6, value=datetime.now())
print(c.number_format)
# 日付の書式を設定できる（月がゼロ埋めされないのはExcelの仕様か、2021/3/13のとき）
c.number_format = 'YYYY/MM/DD H:MM:SS'
# これは文字列になる
ws3.cell(row=7, column=7, value="2021/03/13 13:11:59")

wb2.save(filename = "modified.xlsx")
